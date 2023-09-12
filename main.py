import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename,percent,column):
    wb = xl.load_workbook(filename)
    for sh in range(1, len(wb.sheetnames)+1):
        sheet = wb[f'Sheet{sh}']
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, column)
            correct_value = cell.value * (1+(percent/100))
            correct_price_cell = sheet.cell(row, column)
            correct_price_cell.value = correct_value
            values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=column, max_col=column)

            chart = BarChart()
            chart.add_data(values)
            sheet.add_chart(chart, 't2')

    wb.save(filename)


i = input('No. of files you want to edit: ')
for j in range(int(i)):
    file = input("Enter the  filename you wish to edit:")
    percent = input("By how much percent do you want to change the value: ")
    column = input("Enter the col to be changed: ")
    process_workbook(file,int(percent),int(column))
print("Process Complete!")
